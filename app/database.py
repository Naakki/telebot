import logging
import os
import pathlib
import re
import sqlite3

from openpyxl import load_workbook
from pprint import pprint


# попытка подключиться и разложить на таблицы файл Excel
wb = None
try:
    wb = load_workbook(pathlib.Path('data', 'Combo.xlsx'))

    # Страховка | Телефон | Combo sheets
    telephones = wb['Телефон']
    insurance = wb['Страховка']
    sim = wb['Combo']
except FileNotFoundError:
    logging.info('Oops, where is file?')


def parse_excel_file(sheet, start='A1', end='A1') -> list:
    '''
    Функция парсит данные таблицы и возвращает их в списке
    Входные данные: 
    sheet - Лист таблицы
    start - Начальные координаты ячейки
    end - Последние координаты ячейки (Если не указана строка, до последней)
    '''
        
    if len(end) < 2:
        last = sheet.max_row + 1
        end = str(end) + str(last)

    phones = list()

    for row in sheet[start:end]:
        phone_list = list()

        for cell in row:
            if cell.value == 'Apple' or \
                cell.value == 'Huawei':
                break
            elif cell.value == None:
                pass
            elif cell.value:
                phone_list.append(cell.value)
        
        if (phone_list != []) and (phone_list not in phones):
            logging.info(phone_list) 
            phones.append(phone_list)
    
    return phones


if wb:
    telephones = parse_excel_file(telephones, start='A2', end='F')
    insurance = parse_excel_file(insurance, start='A2', end='F')
    sim_cards = parse_excel_file(sim, end='F')
        

def make_db():
    '''
    Функция формирует БД из данных Excel таблицы, 
    обновляет существующие записи и добавляет новые
    '''

    # Формируется таблица с телефонами
    with sqlite3.connect('data\combo.db') as db:
        logging.info('Connecting database')
        cur = db.cursor()

        cur.execute("""CREATE TABLE IF NOT EXISTS phones(
            brand VARCHAR, 
            model VARCHAR UNIQUE, 
            price MONEY, 
            discount FLOAT)""")

        for phone in telephones:
            if cur.execute("SELECT model FROM phones WHERE model=?", 
                           (phone[1],)).fetchone():
                
                datas = cur.execute("SELECT price, discount FROM phones WHERE model=?", 
                                    (phone[1],)).fetchone()
                
                price, discount = datas
                
                try:
                    if not (int(price) == int(phone[2])) and \
                        not (float(discount) == float(phone[3])):

                        cur.execute("UPDATE phones SET price=?, discount=? WHERE model=?", 
                                    (phone[2], phone[3], phone[1],))
                        db.commit()

                except ValueError as ve:
                    logging.info(ve)

                    if re.fullmatch('\d*', str(phone[2])):
                        cur.execute("UPDATE phones SET price=?, discount=? WHERE model=?", 
                                    (phone[2], phone[3], phone[1],))
                        db.commit()

                        logging.info('Table was updated')
            else:    
                cur.execute("INSERT INTO phones VALUES (?, ?, ?, ?)", phone)
                db.commit()

                logging.info('Table was updated')


        cur.execute("""CREATE TABLE IF NOT EXISTS insurance(
            min_price MONEY, 
            max_price MONEY, 
            insurance_price MONEY)""")

        for row in insurance:
            data = row[0].split('-')
            data = (*data, row[1])
            
            if datas := cur.execute("SELECT * FROM insurance WHERE insurance_price=?", 
                           (row[1],)).fetchone():
                
                min_price, max_price = datas[0], datas[1]

                try:
                    if not (int(min_price) == int(data[0])) and \
                        not (int(max_price) == int(data[1])):

                        cur.execute("UPDATE insurance SET min_price=?, max_price=?, insurance_price=?", data)
                        db.commit()
                except ValueError as ve:
                    logging.info(ve)

            else:
                cur.execute("""INSERT INTO insurance VALUES (?, ?, ?)""", data)
                db.commit()
    
        cur.execute("""CREATE TABLE IF NOT EXISTS sim(
            region VARCAHR,
            sim_price MONEY)""")
        
        for row in sim_cards:
            if datas := cur.execute("SELECT region, sim_price FROM sim WHERE sim_price=?", 
                                    (row[2],)).fetchone():
                
                if not (datas[0] == row[1]) and \
                    not (datas[1] == row[2]):

                    cur.execute("UPDATE sim SET region=?, sim_price=?", (row[1], row[2],))
                    db.commit()

            else:
                cur.execute("INSERT INTO sim VALUES (?, ?)", (row[1], row[2],))
                db.commit()
                


def get_price(model):
    '''
    Выводит запись из таблицы БД    
    '''
    with sqlite3.connect('data\combo.db') as db:
        cur = db.cursor()

        price = cur.execute("""SELECT price FROM phones WHERE model=?""", (model,)).fetchone()[0]

    return price


def get_insurance(model):
    with sqlite3.connect('data\combo.db') as db:
        cur = db.cursor()

        price = cur.execute("""SELECT price FROM phones WHERE model=?""", (model,)).fetchone()[0]
        ins = cur.execute("""SELECT insurance_price FROM insurance 
                                WHERE min_price <= ? and max_price >= ?""", (price, price)).fetchone()
        
        return int(ins[0])
    

def get_combo(model):

    with sqlite3.connect('data\combo.db') as db:
        cur = db.cursor()

        price, discount = cur.execute("""SELECT price, discount FROM phones WHERE model=?""", 
                                      (model,)).fetchone()
        
        ins = get_insurance(model)

        sim = 2150

        return int(price) * (1 - float(discount)) + ins + sim
    

def get_logo(brand):
    with sqlite3.connect('data\combo.db') as db:
        cur = db.cursor()

        logo_bin = cur.execute("""SELECT img FROM logo WHERE brand=?""", (brand,)).fetchone()

    return logo_bin[0]